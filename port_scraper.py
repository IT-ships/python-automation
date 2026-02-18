import os
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime

# ---------------- Folder Setup ----------------
output_folder = "output"
os.makedirs(output_folder, exist_ok=True)

# Use date in filename
today_str = datetime.now().strftime("%Y-%m-%d")
filename = f"combined_arrivals_{today_str}.xlsx"
full_path = os.path.join(output_folder, filename)

# REMOVE old file if exists
if os.path.exists(full_path):
    os.remove(full_path)


# ---------------- Sydney Scraper ----------------
def get_sydney_arrivals():
    BASE_URL = "https://www.portauthoritynsw.com.au/port-operations/sydney-harbour/sydney-harbour-daily-vessel-movements"
    all_data = []
    page = 1

    while True:
        url = f"{BASE_URL}?page={page}"
        response = requests.get(url, timeout=15)
        soup = BeautifulSoup(response.text, "html.parser")

        table = soup.find("table")
        if not table:
            break

        rows = table.find_all("tr")[1:]
        if not rows:
            break

        for row in rows:
            cols = [c.get_text(strip=True) for c in row.find_all("td")]
            if len(cols) >= 8 and cols[2].lower() == "arrival":
                all_data.append({
                    "Port": "Sydney",
                    "Vessel": cols[3],
                    "DateTime": cols[0],
                    "ETA": cols[1],
                    "From": cols[6],
                    "To": cols[7],
                    "Berth": cols[4],
                    "Last Updated": datetime.now().strftime("%Y-%m-%d %H:%M")
                })
        page += 1

    return pd.DataFrame(all_data)

# ---------------- Melbourne Scraper ----------------
def get_melbourne_arrivals():
    URL = "https://ports.vic.gov.au/marine-operations/ship-movements/"
    response = requests.get(URL, timeout=15)
    soup = BeautifulSoup(response.text, "html.parser")

    all_data = []
    sections = soup.find_all("h3")
    for heading in sections:
        title = heading.get_text(strip=True).lower()
        if "arrival" in title:
            table = heading.find_next("table")
            if not table:
                continue
            rows = table.find_all("tr")[1:]
            for row in rows:
                cols = [c.get_text(strip=True) for c in row.find_all("td")]
                if len(cols) >= 4:
                    all_data.append({
                        "Port": "Melbourne",
                        "Category": title,
                        "Vessel": cols[0],
                        "DateTime": cols[1],
                        "From": cols[2],
                        "To": cols[3],
                        "Last Updated": datetime.now().strftime("%Y-%m-%d %H:%M")
                    })
    return pd.DataFrame(all_data)

# ---------------- Main Function ----------------
def main():
    print("Fetching Sydney arrivals...")
    df_sydney = get_sydney_arrivals()
    print(f"Sydney: {len(df_sydney)} arrivals found.")

    print("Fetching Melbourne arrivals...")
    df_melbourne = get_melbourne_arrivals()
    print(f"Melbourne: {len(df_melbourne)} arrivals found.")

    # Write to Excel
    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        if not df_sydney.empty:
            df_sydney.to_excel(writer, sheet_name="Sydney", index=False)
        if not df_melbourne.empty:
            df_melbourne.to_excel(writer, sheet_name="Melbourne", index=False)
            # Add Last Updated sheet
        import openpyxl
        wb = openpyxl.load_workbook(full_path)
        if 'Last Updated' not in wb.sheetnames:
            wb.create_sheet('Last Updated')
        ws = wb['Last Updated']
        ws['A1'] = 'Last Updated'
        ws['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        wb.save(full_path)

    print(f"Excel file saved to: {full_path}")

if __name__ == "__main__":
    main()



